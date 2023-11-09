from io import BytesIO
from sqlite3 import Row
from project import app
from flask import (
    abort,
    render_template, 
    redirect, 
    url_for, 
    request,
    jsonify,
    send_file,
)
from project.utiles.scrape import scraper
from project.config.scraper import LIMIT
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from datetime import datetime
import os

data = []

#route export
@app.route('/export', methods = ['POST'])
def export():
    global data
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add some data to the worksheet
    worksheet['A1'] = 'Company'
    worksheet['B1'] = 'Website'
    worksheet['C1'] = 'Company Linkedin Url'
    worksheet['D1'] = 'Company Phone'
    worksheet['E1'] = 'Company Address'
    worksheet['F1'] = 'Company State'
    worksheet['G1'] = 'Company City'
    worksheet['H1'] = 'Company Postal Code'
    worksheet['I1'] = 'Company Country'
    worksheet['J1'] = 'First Name'
    worksheet['K1'] = 'Last Name'
    worksheet['L1'] = 'Title'
    worksheet['M1'] = 'Email'
    worksheet['N1'] = 'Person Linkedin Url'

    for i in range(len(data)):
        item = data[i]
        if scraper.filter_text != '':
            scraper.add_to_history(item["company"])
        worksheet.cell(row=i + 2, column=1).value = item["company"]
        worksheet.cell(row=i + 2, column=2).value = '=HYPERLINK("{}")'.format(item["website"])
        worksheet.cell(row=i + 2, column=3).value = '=HYPERLINK("{}")'.format(item["linkedin_comp"])
        worksheet.cell(row=i + 2, column=4).value = item["phone"]
        worksheet.cell(row=i + 2, column=5).value = item["address"]
        worksheet.cell(row=i + 2, column=6).value = item["state"]
        worksheet.cell(row=i + 2, column=7).value = item["city"]
        worksheet.cell(row=i + 2, column=8).value = item["code"]
        worksheet.cell(row=i + 2, column=9).value = item["country"]
        worksheet.cell(row=i + 2, column=10).value = item["fname"]
        worksheet.cell(row=i + 2, column=11).value = item["lname"]
        worksheet.cell(row=i + 2, column=12).value = item["title"]
        worksheet.cell(row=i + 2, column=13).value = item["email"]
        worksheet.cell(row=i + 2, column=14).value = '=HYPERLINK("{}")'.format(item["linkedin_pers"])

    # Save the workbook
    try:
        # Set the filename for the workbook
        filename = str(int(datetime.now().timestamp()))+'.xlsx'
        # Set the directory to save the file in
        save_dir = os.path.join(app.root_path, 'csv')
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
        # Create the full path to the file
        filepath = os.path.join(save_dir, filename)
        # Save the workbook to the file
        workbook.save(filepath)
        workbook.close()
        with open(filepath, 'rb') as file:
                file_data = file.read()
                return send_file(BytesIO(file_data), as_attachment=True, mimetype='text/xlsx', download_name=filename)
        # with zipfile.ZipFile(filepath + '.zip', 'w') as zip_file:
        #     zip_file.write(filepath, os.path.basename(filepath))
        #     zip_file.close()
        #     with open(filepath + '.zip', 'rb') as file:
        #         file_data = file.read()
        #         return send_file(BytesIO(file_data), as_attachment=True, mimetype='zip', download_name=filepath + '.zip')
    except Exception as err:
        abort(500, str(err))

#route index
@app.route('/', methods = ['GET', 'POST'])
async def index():
    global data
    if request.method == "POST":
        data = request.get_json()
        loc = data.get("location", None)
        ind = data.get("industry", None)
        job = data.get("job_title", None)
        if loc and ind and job:
            scraper.clear_result_data()
            if scraper.start_scraping(loc, ind, job, LIMIT) == 200:
                data = scraper.get_result_data()
            else:
                scraper.clear_result_data()
        return jsonify({"data": data})
    
    # data.append({
    #     "company": "3D CAM International",
    #     "website": "3d-cam.com",
    #     "linkedin_comp": "http://www.linkedin.com/company/3d-cam-international-corporation",
    #     "phone": "818-773-8777",
    #     "address": "9801 Variel Ave",
    #     "state": "California",
    #     "city": "Los Angeles",
    #     "code": "91311-4317",
    #     "country": "United States",
    #     "fname": "Gary",
    #     "lname": "Vassighi",
    #     "title": "3d Prinitng, Plastic",
    #     "email": "gary@3d-cam.com",
    #     "linkedin_pers": "https://linkedin.com/in/gary-vassighi-931350b9",
    # })
    return render_template('index.html', data=data)
